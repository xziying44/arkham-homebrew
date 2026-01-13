import re
import openpyxl
from openpyxl.cell.cell import Cell
from copy import copy
import os


def replace_quotes_in_excel(input_path: str, output_path: str):
    """
    读取一个Excel文件，将所有单元格中的英文引号替换为中文引号，并保存为新文件。
    - 将 "text" 替换为 “text”
    - 将 'text' 替换为 ‘text’

    Args:
        input_path (str): 源Excel文件的路径.
        output_path (str): 处理后要保存的新Excel文件的路径.
    """
    # --- 1. 参数校验和文件加载 ---
    if not os.path.exists(input_path):
        print(f"错误：输入文件不存在 -> {input_path}")
        return

    print(f"正在加载工作簿: {input_path} ...")
    try:
        source_wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        print(f"无法加载Excel文件，请确保文件格式正确且未损坏。错误：{e}")
        return

    dest_wb = openpyxl.Workbook()
    # 移除默认创建的Sheet，以便后续完全复制源文件的结构
    if 'Sheet' in dest_wb.sheetnames:
        dest_wb.remove(dest_wb['Sheet'])

    # --- 2. 定义正则表达式 ---
    # 匹配英文双引号并捕获其中的内容
    double_quote_pattern = re.compile(r'"(.*?)"')
    # 匹配英文单引号并捕获其中的内容
    single_quote_pattern = re.compile(r"'(.*?)'")

    # --- 3. 遍历和处理 ---
    for sheet_name in source_wb.sheetnames:
        source_sheet = source_wb[sheet_name]
        dest_sheet = dest_wb.create_sheet(title=sheet_name)

        print(f"正在处理工作表: '{sheet_name}' ...")

        # 复制工作表级别的设置，如打印选项、页面边距等
        dest_sheet.sheet_properties.tabColor = source_sheet.sheet_properties.tabColor
        # ... 可以根据需要复制更多工作表级别的属性

        # 遍历所有行和单元格
        for row in source_sheet.iter_rows():
            for source_cell in row:
                dest_cell = dest_sheet.cell(row=source_cell.row, column=source_cell.column)

                # 复制值并进行处理
                if isinstance(source_cell.value, str):
                    original_value = source_cell.value
                    # 首先替换双引号
                    new_value = double_quote_pattern.sub(r'“\1”', original_value)
                    # 然后在结果上替换单引号
                    new_value = single_quote_pattern.sub(r'‘\1’', new_value)
                    dest_cell.value = new_value
                else:
                    # 如果不是字符串（数字、日期、None等），直接赋值
                    dest_cell.value = source_cell.value

                # 复制样式（保持格式）
                if source_cell.has_style:
                    dest_cell.font = copy(source_cell.font)
                    dest_cell.border = copy(source_cell.border)
                    dest_cell.fill = copy(source_cell.fill)
                    dest_cell.number_format = copy(source_cell.number_format)
                    dest_cell.protection = copy(source_cell.protection)
                    dest_cell.alignment = copy(source_cell.alignment)

    # --- 4. 保存文件 ---
    try:
        print(f"\n正在保存到新文件: {output_path} ...")
        dest_wb.save(output_path)
        print("处理完成！")
    except Exception as e:
        print(f"保存文件时出错，请检查路径和权限。错误：{e}")


if __name__ == '__main__':
    # --- 使用说明 ---
    # 1. 将你的源Excel文件命名为 "source.xlsx" 并放在此脚本相同的目录下。
    #    或者，直接修改下面的 input_file 变量为你的文件路径。
    input_file = "ahlcg简中信息录入 (25).xlsx"

    # 2. 定义输出文件名。
    output_file = "processed_output.xlsx"

    # 3. 运行此脚本。
    replace_quotes_in_excel(input_file, output_file)

